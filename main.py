import openpyxl
import os, datetime
import re
import copy

current_folder = input("Введите путь к папке: ")
print(f'Указана папка: {current_folder}')
# Обходим папку
folder_walk = os.walk(current_folder, onerror=None, followlinks=False)

# Создаем новый файл
new_excel_file = openpyxl.Workbook()
ws_new_excel_file = new_excel_file.active
ws_new_excel_file.title = 'Export'
folder_walk_list = []

new_excel_log_file = openpyxl.Workbook()
ws_new_excel_log_file = new_excel_log_file.active
ws_new_excel_log_file.title = 'Exception'

for i in folder_walk:
    folder_walk_list.append(i)
    print(i)

title_row = ['№', '№ поз.', '№ ЛСР', '№ ОСР', 'Специальность', 'Основание ЛСР', 'Тип затрат', 'Обоснование позиции',
             'Наименование', 'Кол-во', 'Стоимость за единицу (в базисном уровне)',
             'Общая стоимость (в базисном уровне)',
             'Стоимость за единицу (в текущем уровне цен  без НДС)',
             'Общая стоимость (в текущем уровне цен  без НДС)',
             'Связанная смета', 'ОБ', 'СМР',
             'Стройка']

title_row_log = ['№', 'Exception', 'Link']
ws_new_excel_log_file.append(title_row_log)

ws_new_excel_file.append(title_row)
count = 1

for folder in folder_walk_list:
    print(f'Папка: {folder[0]}')
    for current_file in folder[2]:
        try:
            file_path_full = os.path.join(folder[0], current_file)  # Составляем полный путь к файлу
            wb = openpyxl.load_workbook(filename=file_path_full, data_only=True)
            print(f'Файл: {current_file}')

            # Получение индексов для оборудования и материалов
            # Определение таблицы ЛСР
            for i in range(len(wb.sheetnames)):
                if re.search('ЛС', str(wb.sheetnames[i]).upper()):
                    worksheet_for_kf_name = str(wb.sheetnames[i])
                else:
                    worksheet_for_kf_name = str(wb.sheetnames[1])

            worksheet_for_kf_name = str(wb.sheetnames[1])
            print(worksheet_for_kf_name)
            worksheet_for_kf = wb[worksheet_for_kf_name]
            first_list_len = worksheet_for_kf.max_row + 1
            kf_equipment = 0
            kf_equipment_count = first_list_len
            kf_smr_count = first_list_len
            kf_smr = 0


            while kf_smr == 0 and kf_smr_count > 0:
                cell_kf = f'A{kf_smr_count}'
                if 'с учётом индекса пересчёта на СМР:' in str(worksheet_for_kf[cell_kf].value):
                    kf_smr_str = worksheet_for_kf[cell_kf].value
                    kf_smr_re = str(re.findall(r'\d{1,4},\d{1,4}', kf_smr_str)[0])
                    kf_smr = float('.'.join(kf_smr_re.split(',')))
                    # kf_smr = float(kf_smr_re)
                kf_smr_count -= 1
            while kf_equipment == 0 and kf_equipment_count > 0:
                cell_kf = f'A{kf_equipment_count}'
                if 'с учётом индекса пересчёта на оборудование:' in str(worksheet_for_kf[cell_kf].value):
                    kf_equipment_str = worksheet_for_kf[cell_kf].value
                    kf_equipment_re = str(re.findall(r'\d{1,4},\d{1,4}', kf_equipment_str)[0])
                    kf_equipment = float('.'.join(kf_equipment_re.split(',')))
                    # kf_equipment = float(kf_equipment_re)
                kf_equipment_count -= 1

            kf_equipment_global = copy.copy(kf_equipment)
            kf_smr_global = copy.copy(kf_smr)
            print(f'Коэффициент оборудования kf_equipment = {kf_equipment}')
            print(f'Коэффициент СМР kf_smr = {kf_smr}')

            # Работа с таблицей Source

            hyperlink = f'{file_path_full}'

            worksheet = wb["Source"]  # Назначаем лист
            book_len = worksheet.max_row + 1
            print(f'Общее количество строк файла: {book_len}')
            object_name = worksheet['G4'].value  # Название объекта
            print(f'Объект: {object_name}')
            number_lsr_list = str(worksheet['F12'].value).split(' ')
            # number_osr = str(number_lsr_list[1])[1:-1]
            number_lsr = str(number_lsr_list[0])
            number_osr_list = number_lsr.split('-')
            number_osr = f'{number_osr_list[1]}-{number_osr_list[2]}'
            spec = ''
            try:
                spec_temp = int(str(number_lsr_list[0])[-3:][1:-1])
                if spec_temp == 1:
                    spec = 'СТР'
                elif spec_temp == 2:
                    spec = 'САН'
                elif spec_temp == 3:
                    spec = 'ТХ'
                elif spec_temp == 4:
                    spec = 'ЭЛ'
            except Exception as e:
                print(f'Ошибка: {e}')

            for i in range(1, book_len):
                cell = f'EG{i}'
                if kf_equipment_global == 0:
                    if worksheet[f'BC{i}'].value:
                        kf_equipment = float(worksheet[f'BC{i}'].value)
                        # print('kf_equipment', type(worksheet[f'BC{i}'].value), worksheet[f'BC{i}'].value)
                if kf_smr_global == 0:
                    if worksheet[f'BC{i}'].value:
                        kf_smr = float(worksheet[f'BC{i}'].value)
                        # print('kf_smr', type(worksheet[f'BC{i}'].value), worksheet[f'BC{i}'].value)
                value = str(worksheet[cell].value).upper()
                equipment_flag = re.search('ОБОРУДОВАНИЕ', value)
                material_flag = re.search('МАТЕРИАЛ', value)
                zatrati_flag = re.search('ЗАТРАТЫ',
                                         str(worksheet[f'G{i}'].value).upper())  # Проверка по слову "Затраты на..."
                # print(f'Затрата {zatrati_flag}, {str(worksheet[f"G{i}"].value).upper()}')
                obosnovanie_position_cell = f'F{i}'
                obosnovanie_position_cell_fer_flag = re.search('ФЕР-',
                                                               str(worksheet[f'F{i}'].value))
                obosnovanie_position_cell_ferm_flag = re.search('ФЕРм-',
                                                                str(worksheet[f'F{i}'].value))
                obosnovanie_position = worksheet[obosnovanie_position_cell].value
                if str(worksheet[f'E{i}'].value).upper() == '':
                    print(f'Пропущено: строка {i}, столбец E: "{str(worksheet[f"E{i}"].value).upper()}"')
                elif zatrati_flag:
                    print(
                        f'Пропущено: строка {i}, zatrati_flag = {zatrati_flag}, столбец G: "{str(worksheet[f"G{i}"].value).upper()}"')
                elif obosnovanie_position_cell_fer_flag and obosnovanie_position_cell_ferm_flag:
                    print(
                        f'Пропущено: строка {i}, ferm_flag = {obosnovanie_position_cell_ferm_flag}, fer_flag = {obosnovanie_position_cell_fer_flag}, столбец BJ: "{str(worksheet[f"BJ{i}"].value)}"')
                else:
                    if equipment_flag or material_flag:
                        position_lsr_cell = f'E{i}'
                        position_lsr = worksheet[position_lsr_cell].value
                        obosnovanie_lsr_cell = f'CN{i}'
                        obosnovanie_lsr = str(worksheet[obosnovanie_lsr_cell].value)

                        name_cell = f'G{i}'
                        name_value = worksheet[name_cell].value
                        amount_cell = f'I{i}'
                        amount_value = worksheet[amount_cell].value

                        price_per_one_basis = float(worksheet[f'AC{i}'].value)
                        price_total_basis = float(worksheet[f'O{i}'].value)

                        if equipment_flag:
                            type_zatrat = 'ОБ'
                            price_per_one_now = kf_equipment * float(worksheet[f'AC{i}'].value) * 1.012 * 1.03
                            # price_per_one_now = 6.16 * float(worksheet[f'AC{i}'].value) * 1.012 * 1.03
                            price_total_now = price_per_one_now * amount_value
                        elif material_flag:
                            type_zatrat = 'МАТ'
                            if re.search('ФССЦ', obosnovanie_position):
                                price_per_one_now = kf_smr * float(worksheet[f'AC{i}'].value)
                                # price_per_one_now = 13.16 * float(worksheet[f'AC{i}'].value)
                                price_total_now = price_per_one_now * amount_value
                            else:
                                price_per_one_now = kf_smr * float(worksheet[f'AC{i}'].value) * 1.02
                                # price_per_one_now = 13.16 * float(worksheet[f'AC{i}'].value) * 1.2
                                price_total_now = price_per_one_now * amount_value
                        else:
                            type_zatrat = ''
                            price_per_one_now = ''
                            price_total_now = ''

                        row = [count, position_lsr, number_lsr, number_osr, spec, obosnovanie_lsr, type_zatrat,
                               obosnovanie_position, name_value, amount_value, price_per_one_basis, price_total_basis,
                               price_per_one_now, price_total_now, hyperlink, kf_equipment, kf_smr, object_name]

                        ws_new_excel_file.append(row)
                        print(f'Row №{i}: {row}')
                        count += 1
            print(f"Файл {current_file} отработан")

            # print(excel.head())
        except Exception as e:
            print(f'Ошибка выполнения файла {current_file}. {e}')
            file_path_full = os.path.join(folder[0], current_file)
            hyperlink = f'{file_path_full}'
            row_exception = [count, str(e), hyperlink]
            ws_new_excel_log_file.append(row_exception)

        print('-----------------------')

now_time = f'{datetime.datetime.now().year}-{datetime.datetime.now().month}-{datetime.datetime.now().day}_{datetime.datetime.now().hour}-{datetime.datetime.now().minute}-{datetime.datetime.now().second}'
file_export_name = f'export-{now_time}.xlsx'
file_export_log_name = f'export-log-{now_time}.xlsx'
file_export_path_full = os.path.join(current_folder, file_export_name)
file_export_path_full_log = os.path.join(current_folder, file_export_log_name)
print(f'Файл {file_export_path_full} готов.')

new_excel_file.save(file_export_path_full)
new_excel_log_file.save(file_export_path_full_log)
